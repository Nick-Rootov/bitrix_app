from openpyxl import load_workbook
from typing import List, Dict
from .base_importer import BaseImporter


class XLSXImporter(BaseImporter): #Реализация импорта данных из XLSX файла (Excel)

    def import_data(self, file_path: str) -> List[Dict]:

        try:
            wb = load_workbook(filename=file_path)
            ws = wb.active

            # Получаем заголовки из первой строки
            headers = [cell.value for cell in ws[1]]

            # Собираем данные
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):  # Пропускаем пустые строки
                    data.append(dict(zip(headers, row)))

            if not data:
                raise ValueError("XLSX файл не содержит данных")

            return data

        except Exception as e:
            raise Exception(f"Ошибка чтения XLSX файла: {str(e)}")